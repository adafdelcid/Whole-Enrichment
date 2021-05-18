# Author: Ada Del Cid
# GitHub: @adafdelcid
# April 2021

import math
from openpyxl import Workbook
import pandas as pd
import numpy as np
import openpyxl

pd.options.mode.chained_assignment = None


def run_enrichment_analysis(destination_folder, formulations_sheet, csv_filepath, sorted_cells, number_naked_bcs,
                            x_percent, sample_numbers):
    """
    run_enrichment_analysis : driver function, it uses all other functions to create enrichment analysis
        inputs:
                destination_folder : user specified path to the folder where the user wants the excel file created to be
                                     saved
                data_sheet : user specified file path to excel spreadsheet with formulation sheet only
                sorted_cells : user specified list of cells that were sorted
    """
    # order list of sorted cells alphabetically
    sorted_cells.sort()

    # create excel destination file
    destination_file = create_excel_spreadsheet(destination_folder)

    # Read formulation sheet and save as dataframe
    df_formulations = create_df_formulation_sheet(formulations_sheet)

    # list of components
    list_components = df_formulations.columns.tolist()
    list_components.pop(0)
    list_components.pop(0)
    list_components.pop()

    # Read CSV file and save as dataframe
    df_norm_counts = create_df_norm_counts(csv_filepath, sample_numbers)

    # Merge dataframes
    df_merged = merge_formulations_and_norm_counts(df_formulations, df_norm_counts, destination_file,
                                                   "Formulations + norm counts")

    # get ordered list of all samples
    d_samples_by_cell_type = divide_samples_by_cell_type(df_merged, sorted_cells)
    # print(d_samples_by_cell_type)

    # divide samples by cell types
    dict_df_avg_cell_type = df_cell_types(df_merged, d_samples_by_cell_type)

    list_organs = get_list_organs(sorted_cells)
    # print(list_organs)
    dict_df_organs = df_by_organs(df_merged, sorted_cells, dict_df_avg_cell_type, list_organs)
    df_overall = get_df_overall(dict_df_organs, df_formulations)

    # sort normalized counts by overall average
    df_sorted = sort_norm_counts(df_overall, -1)

    dict_components = get_lists_of_components(df_formulations, list_components, number_naked_bcs)

    df_top, df_bottom = df_top_and_bottom(df_sorted, x_percent, number_naked_bcs)

    # keep this code: DO NOT DELETE
    create_all_sheet(destination_file, dict_df_organs, df_overall, df_top, df_bottom, dict_components)
    create_cell_type_sheets(destination_file, df_formulations, dict_df_avg_cell_type, dict_components,
                            d_samples_by_cell_type, x_percent, number_naked_bcs)

    d_organ_sheet_columns = get_column_names_organ_sheets(d_samples_by_cell_type, list_organs, sample_numbers)

    create_organ_sheet(destination_file, df_formulations, df_norm_counts, dict_components, d_organ_sheet_columns,
                       x_percent, number_naked_bcs)


def create_organ_sheet(destination_file, df_formulations, df_norm_counts, dict_components, d_organ_sheet_columns,
                       x_percent, number_naked_bcs):
    positions_dict = {}
    with pd.ExcelWriter(destination_file, engine="openpyxl", mode="a") \
            as writer:  # pylint: disable=abstract-class-instantiated
        for organ in d_organ_sheet_columns:
            current_col = 0  # variable to place formulation enrichments by mole ratio

            positions_organ_samples = {}
            for sample_num in d_organ_sheet_columns[organ]:
                df_mouse = df_norm_counts[d_organ_sheet_columns[organ][sample_num]]
                avg = df_mouse.mean(axis=1)

                # sorted averaged cell type dataframe
                temp_df = pd.concat([df_formulations, df_mouse], axis=1)
                avg_col_name = sample_num + "-AVG"
                temp_df[avg_col_name] = avg

                df_sorted_avg = sort_norm_counts(temp_df, -1)  # sort by avg

                df_sorted_avg.to_excel(writer, sheet_name=organ, startrow=0, startcol=current_col, index=False)

                # top & bottom
                df_top_avg, df_bottom_avg = df_top_and_bottom(df_sorted_avg, x_percent, number_naked_bcs)

                d_df_avg_components_top, d_df_avg_components_bottom = top_bottom_enrichment(df_sorted_avg,
                                                                                            dict_components, df_top_avg,
                                                                                            df_bottom_avg)
                # total
                dict_df_component_enrichments = get_overall_enrichment(df_sorted_avg, dict_components)

                d_df_component_net_enrichment, d_df_enrichment_factors_top, d_df_enrichment_factors_bottom = \
                    net_enrichment_factor(dict_df_component_enrichments, d_df_avg_components_top,
                                          d_df_avg_components_bottom, sort_by=avg_col_name)

                current_col += len(temp_df.columns) + 1

                df_top_avg.to_excel(writer, sheet_name=organ, startrow=0, startcol=current_col, index=False)

                current_row = len(df_top_avg) + 2

                df_bottom_avg.to_excel(writer, sheet_name=organ, startrow=current_row, startcol=current_col,
                                       index=False)

                current_col += len(df_top_avg.columns) + 1
                current_row = 1
                start_col_enrichment_tables = current_col + 1

                positions_organ_samples[sample_num] = start_col_enrichment_tables

                for component in dict_df_component_enrichments:
                    dict_df_component_enrichments[component].to_excel(writer, sheet_name=organ, startrow=current_row,
                                                                      startcol=current_col, index=False)
                    d_df_avg_components_top[component].to_excel(writer, sheet_name=organ, startrow=current_row,
                                                                startcol=current_col + 4, index=False)
                    d_df_enrichment_factors_top[component].to_excel(writer, sheet_name=organ, startrow=current_row,
                                                                    startcol=current_col + 8, index=False)
                    d_df_avg_components_bottom[component].to_excel(writer, sheet_name=organ, startrow=current_row,
                                                                   startcol=current_col + 11, index=False)
                    d_df_enrichment_factors_bottom[component].to_excel(writer, sheet_name=organ, startrow=current_row,
                                                                       startcol=current_col + 15, index=False)
                    d_df_component_net_enrichment[component].to_excel(writer, sheet_name=organ, startrow=current_row,
                                                                      startcol=current_col + 18, index=False)

                    current_row += len(dict_df_component_enrichments[component]) + 2

                current_col += 21

            positions_dict[organ] = positions_organ_samples

    xfile = openpyxl.load_workbook(destination_file)

    for organ in d_organ_sheet_columns:
        sheet = xfile[organ]
        start_col_enrichment_tables = positions_dict[organ]
        mouse_samples = d_organ_sheet_columns[organ]
        for sample in mouse_samples:
            sheet.cell(row=1, column=start_col_enrichment_tables[sample]).value = "Total"
            sheet.cell(row=1, column=start_col_enrichment_tables[sample] + 4).value = "Top"
            sheet.cell(row=1, column=start_col_enrichment_tables[sample] + 8).value = "Enrichment-Top"
            sheet.cell(row=1, column=start_col_enrichment_tables[sample] + 11).value = "Bottom"
            sheet.cell(row=1, column=start_col_enrichment_tables[sample] + 15).value = "Depletion-Bottom"
            sheet.cell(row=1, column=start_col_enrichment_tables[sample] + 18).value = "Net Enrichment Factor"
    xfile.save(destination_file)


def get_column_names_organ_sheets(d_samples_by_cell_type, list_organs, sample_numbers):
    d_organs_d_cell_types_samples = get_dict_organs_by_cell_type(d_samples_by_cell_type, list_organs)
    d_organ_sheet_columns = {}

    for organ in list_organs:
        cell_type_samples = d_organs_d_cell_types_samples[organ]
        temp_dict = {}
        for sample_num in sample_numbers:
            temp_list = []
            for cell_type in cell_type_samples:
                for sample in cell_type_samples[cell_type]:
                    if sample_num in sample:
                        temp_list.append(sample)
            if len(temp_list) != 0:
                temp_dict[sample_num] = temp_list

        d_organ_sheet_columns[organ] = temp_dict

    return d_organ_sheet_columns


def get_dict_organs_by_cell_type(d_samples_by_cell_type, list_organs):
    d_organs_by_cell_type = {}

    for organ in list_organs:
        d_organs_by_cell_type[organ] = {}
        for sample_cell_type in d_samples_by_cell_type:
            if sample_cell_type[0] == organ:
                d_organs_by_cell_type[organ][sample_cell_type] = d_samples_by_cell_type[sample_cell_type]

    return d_organs_by_cell_type


def create_cell_type_sheets(destination_file, df_formulations, dict_df_avg_cell_type, dict_components,
                            d_samples_by_cell_type, x_percent, number_naked_bcs):
    positions_dict = {}
    with pd.ExcelWriter(destination_file, engine="openpyxl", mode="a") \
            as writer:  # pylint: disable=abstract-class-instantiated
        for cell_type in dict_df_avg_cell_type:
            current_col = 0  # variable to place formulation enrichments by mole ratio

            # sorted averaged cell type dataframe
            temp_df = pd.concat([df_formulations, dict_df_avg_cell_type[cell_type]], axis=1)
            df_sorted_avg = sort_norm_counts(temp_df, -2)  # sort by avg
            df_sorted_avg.to_excel(writer, sheet_name=cell_type, startrow=0, startcol=current_col, index=False)

            # top & bottom
            df_top_avg, df_bottom_avg = df_top_and_bottom(df_sorted_avg, x_percent, number_naked_bcs)

            d_df_avg_components_top, d_df_avg_components_bottom = top_bottom_enrichment(df_sorted_avg, dict_components,
                                                                                        df_top_avg, df_bottom_avg)

            # total
            dict_df_component_enrichments = get_overall_enrichment(df_sorted_avg, dict_components)

            d_df_component_net_enrichment, d_df_enrichment_factors_top, d_df_enrichment_factors_bottom = \
                net_enrichment_factor(dict_df_component_enrichments, d_df_avg_components_top,
                                      d_df_avg_components_bottom, sort_by=cell_type)

            current_col += len(temp_df.columns) + 1

            df_top_avg.to_excel(writer, sheet_name=cell_type, startrow=0, startcol=current_col, index=False)

            current_row = len(df_top_avg) + 2

            df_bottom_avg.to_excel(writer, sheet_name=cell_type, startrow=current_row, startcol=current_col,
                                   index=False)

            current_col += len(df_top_avg.columns) + 1
            current_row = 1
            start_col_enrichment_tables = current_col + 1
            positions_dict[cell_type] = start_col_enrichment_tables

            for component in dict_df_component_enrichments:
                dict_df_component_enrichments[component].to_excel(writer, sheet_name=cell_type, startrow=current_row,
                                                                  startcol=current_col, index=False)
                d_df_avg_components_top[component].to_excel(writer, sheet_name=cell_type, startrow=current_row,
                                                            startcol=current_col + 4, index=False)
                d_df_enrichment_factors_top[component].to_excel(writer, sheet_name=cell_type, startrow=current_row,
                                                                startcol=current_col + 8, index=False)
                d_df_avg_components_bottom[component].to_excel(writer, sheet_name=cell_type, startrow=current_row,
                                                               startcol=current_col + 11, index=False)
                d_df_enrichment_factors_bottom[component].to_excel(writer, sheet_name=cell_type, startrow=current_row,
                                                                   startcol=current_col + 15, index=False)
                d_df_component_net_enrichment[component].to_excel(writer, sheet_name=cell_type, startrow=current_row,
                                                                  startcol=current_col + 18, index=False)

                current_row += len(dict_df_component_enrichments[component]) + 2

            current_col += 21

            for sample_cell_type in d_samples_by_cell_type[cell_type]:
                df_sample = dict_df_avg_cell_type[cell_type][sample_cell_type]
                temp_df = pd.concat([df_formulations, df_sample], axis=1)
                df_sorted = sort_norm_counts(temp_df, -1)  # sort by sample

                df_sorted.to_excel(writer, sheet_name=cell_type, startrow=0, startcol=current_col, index=False)

                current_col += len(df_sorted.columns) + 1

                # top & bottom
                df_top_avg, df_bottom_avg = df_top_and_bottom(df_sorted, x_percent, number_naked_bcs)

                d_df_avg_components_top, d_df_avg_components_bottom = top_bottom_enrichment(df_sorted,
                                                                                            dict_components, df_top_avg,
                                                                                            df_bottom_avg)

                # total
                dict_df_component_enrichments = get_overall_enrichment(df_sorted, dict_components)

                d_df_component_net_enrichment, d_df_enrichment_factors_top, d_df_enrichment_factors_bottom = \
                    net_enrichment_factor(dict_df_component_enrichments, d_df_avg_components_top,
                                          d_df_avg_components_bottom, sort_by=sample_cell_type)

                df_top_avg.to_excel(writer, sheet_name=cell_type, startrow=0, startcol=current_col, index=False)

                current_row = len(df_top_avg) + 2

                df_bottom_avg.to_excel(writer, sheet_name=cell_type, startrow=current_row, startcol=current_col,
                                       index=False)

                current_col += len(df_top_avg.columns) + 1
                current_row = 1

                for component in dict_df_component_enrichments:
                    dict_df_component_enrichments[component].to_excel(writer, sheet_name=cell_type,
                                                                      startrow=current_row,
                                                                      startcol=current_col, index=False)
                    d_df_avg_components_top[component].to_excel(writer, sheet_name=cell_type, startrow=current_row,
                                                                startcol=current_col + 4, index=False)
                    d_df_enrichment_factors_top[component].to_excel(writer, sheet_name=cell_type, startrow=current_row,
                                                                    startcol=current_col + 8, index=False)
                    d_df_avg_components_bottom[component].to_excel(writer, sheet_name=cell_type, startrow=current_row,
                                                                   startcol=current_col + 11, index=False)
                    d_df_enrichment_factors_bottom[component].to_excel(writer, sheet_name=cell_type,
                                                                       startrow=current_row,
                                                                       startcol=current_col + 15, index=False)
                    d_df_component_net_enrichment[component].to_excel(writer, sheet_name=cell_type,
                                                                      startrow=current_row,
                                                                      startcol=current_col + 18, index=False)

                    current_row += len(dict_df_component_enrichments[component]) + 2
                current_col += 21
    xfile = openpyxl.load_workbook(destination_file)
    for cell_type in dict_df_avg_cell_type:
        sheet = xfile[cell_type]
        start_col_enrichment_tables = positions_dict[cell_type]
        for i in range(len(d_samples_by_cell_type[cell_type]) + 1):
            sheet.cell(row=1, column=start_col_enrichment_tables).value = "Total"
            sheet.cell(row=1, column=start_col_enrichment_tables + 4).value = "Top"
            sheet.cell(row=1, column=start_col_enrichment_tables + 8).value = "Enrichment-Top"
            sheet.cell(row=1, column=start_col_enrichment_tables + 11).value = "Bottom"
            sheet.cell(row=1, column=start_col_enrichment_tables + 15).value = "Depletion-Bottom"
            sheet.cell(row=1, column=start_col_enrichment_tables + 18).value = "Net Enrichment Factor"
            start_col_enrichment_tables += 47
    xfile.save(destination_file)


def create_all_sheet(destination_file, dict_df_organs, df_overall, df_top, df_bottom, dict_components):
    # top & bottom
    d_df_components_top, d_df_components_bottom = top_bottom_enrichment(df_overall, dict_components, df_top, df_bottom)

    # total
    dict_df_component_enrichments = get_overall_enrichment(df_overall, dict_components)

    d_df_component_net_enrichment, d_df_enrichment_factors_top, d_df_enrichment_factors_bottom = \
        net_enrichment_factor(dict_df_component_enrichments, d_df_components_top, d_df_components_bottom,
                              sort_by="Overall-AVG")

    with pd.ExcelWriter(destination_file, engine="openpyxl", mode="a") \
            as writer:  # pylint: disable=abstract-class-instantiated

        current_col = 0  # variable to place formulation enrichments by mole ratio
        my_sheet_name = "All"

        for component in dict_df_organs:
            dict_df_organs[component].to_excel(writer, sheet_name=my_sheet_name, startrow=0, startcol=current_col,
                                               index=False)

            current_col += len(dict_df_organs[component].columns) + 1

        df_overall.to_excel(writer, sheet_name=my_sheet_name, startrow=0, startcol=current_col, index=False)

        current_col += len(df_overall.columns) + 1

        df_top.to_excel(writer, sheet_name=my_sheet_name, startrow=0, startcol=current_col, index=False)

        current_row = len(df_top) + 2

        df_bottom.to_excel(writer, sheet_name=my_sheet_name, startrow=current_row, startcol=current_col, index=False)

        current_col += len(df_top.columns) + 1
        current_row = 1
        start_col_enrichment_tables = current_col + 1

        for component in dict_df_component_enrichments:

            dict_df_component_enrichments[component].to_excel(writer, sheet_name=my_sheet_name, startrow=current_row,
                                                              startcol=current_col, index=False)
            d_df_components_top[component].to_excel(writer, sheet_name=my_sheet_name, startrow=current_row,
                                                    startcol=current_col + 4, index=False)
            d_df_enrichment_factors_top[component].to_excel(writer, sheet_name=my_sheet_name, startrow=current_row,
                                                            startcol=current_col + 8, index=False)
            d_df_components_bottom[component].to_excel(writer, sheet_name=my_sheet_name, startrow=current_row,
                                                       startcol=current_col + 11, index=False)
            d_df_enrichment_factors_bottom[component].to_excel(writer, sheet_name=my_sheet_name, startrow=current_row,
                                                               startcol=current_col + 15, index=False)
            d_df_component_net_enrichment[component].to_excel(writer, sheet_name=my_sheet_name, startrow=current_row,
                                                              startcol=current_col + 18, index=False)

            current_row += len(dict_df_component_enrichments[component]) + 2

    xfile = openpyxl.load_workbook(destination_file)
    sheet = xfile[my_sheet_name]
    sheet.cell(row=1, column=start_col_enrichment_tables).value = "Total"
    sheet.cell(row=1, column=start_col_enrichment_tables + 4).value = "Top"
    sheet.cell(row=1, column=start_col_enrichment_tables + 8).value = "Enrichment-Top"
    sheet.cell(row=1, column=start_col_enrichment_tables + 11).value = "Bottom"
    sheet.cell(row=1, column=start_col_enrichment_tables + 15).value = "Depletion-Bottom"
    sheet.cell(row=1, column=start_col_enrichment_tables + 18).value = "Net Enrichment Factor"
    xfile.save(destination_file)


def dict_list_to_dict_df(dict_list, sort_by="AVG"):
    """
    dict_list_to_dict_df: converts dictionary with lists to dictionary with dataframes
        inputs:
            dict_list: dictionary containing lists
            sort_by : user specified cell type to sort by, default is "AVG"
        output:
            dict_df : dictionary with dataframes
    """

    dict_df = {}
    for component in dict_list:
        np_temporary = np.array(dict_list[component])
        dict_df[component] = pd.DataFrame(data=np_temporary, columns=[component, sort_by])

    return dict_df


def net_enrichment_factor(dict_df_component_enrichments, d_df_components_top, d_df_components_bottom, sort_by="AVG"):
    """
    net_enrichment_factor: creates dataframes for best and worst performing LNPs, counts and their formulations
        inputs:
            d_df_components_averaged : dictionary with all dataframes of all enrichment calculations of df_averaged
            d_df_components_top : dictionary with all dataframes of all enrichment calculations of df_top
            d_df_components_bottom : dictionary with all dataframes of all enrichment calculations of df_bottom
            sort_by : user specified cell type to sort by, default is "AVG"
        output:
            d_df_component_net_enrichment : dictionary with dataframes of net enrichment factors by component type or
                                             mole ratio
            d_raw_enrichment_factors_top: dictionary with dataframes of raw enrichment of top performing LNPs
            d_raw_enrichment_factors_bottom: dictionary with dataframes of raw enrichment of bottom performing LNPs
    """

    dict_component_net_enrichment_factor = {}

    d_raw_enrichment_factors_top = raw_enrichment_factor(dict_df_component_enrichments, d_df_components_top)
    d_raw_enrichment_factors_bottom = raw_enrichment_factor(dict_df_component_enrichments, d_df_components_bottom)

    for component in d_raw_enrichment_factors_top:
        temporary_list = []
        for index in range(len(d_raw_enrichment_factors_top[component])):
            enrichment_factor_row_top = d_raw_enrichment_factors_top[component][index]
            enrichment_factor_row_bottom = d_raw_enrichment_factors_bottom[component][index]
            item = [enrichment_factor_row_top[0], round(enrichment_factor_row_top[1] - enrichment_factor_row_bottom[1],
                                                        9)]
            temporary_list.append(item)

        dict_component_net_enrichment_factor[component] = temporary_list

    d_df_component_net_enrichment = {}

    for component in dict_component_net_enrichment_factor:
        np_temporary = np.array(dict_component_net_enrichment_factor[component])
        d_df_component_net_enrichment[component] = pd.DataFrame(data=np_temporary, columns=[component, sort_by])

    dict_df_raw_enrichment_top = dict_list_to_dict_df(d_raw_enrichment_factors_top, sort_by)
    dict_df_raw_enrichment_bottom = dict_list_to_dict_df(d_raw_enrichment_factors_bottom, sort_by)

    return d_df_component_net_enrichment, dict_df_raw_enrichment_top, dict_df_raw_enrichment_bottom


def raw_enrichment_factor(dict_df_component_enrichments, d_df_components_top_bottom):
    """
    raw_enrichment_factor: creates dataframes for best and worst performing LNPs, counts and their
                            formulations
        inputs:
            d_df_components_averaged : dictionary with all dataframes of all enrichment
                                        calculations of df_overall
            d_df_components_top_bottom : dictionary with all dataframes of all enrichment
                                        calculations of df_top_bottom_sort_by
        output:
            dict_raw_enrichment_factors : dictionary with lists of all raw enrichment factors
    """

    dict_components_averaged = {}
    dict_components_top_bottom = {}
    dict_raw_enrichment_factors = {}

    for component in dict_df_component_enrichments:
        dict_components_averaged[component] = dict_df_component_enrichments[component].values.tolist()
        dict_components_top_bottom[component] = d_df_components_top_bottom[component].values.tolist()

        temporary_list = []
        for index in range(len(dict_components_averaged[component])):
            overall_row = dict_components_averaged[component][index]
            top_bottom_row = dict_components_top_bottom[component][index]
            item = [overall_row[0], round(float(top_bottom_row[2]) / float(overall_row[2]), 9)]
            temporary_list.append(item)

        dict_raw_enrichment_factors[component] = temporary_list

    return dict_raw_enrichment_factors


def get_overall_enrichment(df_overall, dict_components):
    try:
        for component in dict_components:
            temp_list = dict_components[component]
            while "TOTAL" in temp_list:
                temp_list.remove("TOTAL")
    except ValueError:
        pass
    dict_df_component_enrichment = {}

    for component in dict_components:
        dict_df_component_enrichment[component] = calculate_enrichment(component, dict_components[component],
                                                                       df_overall)
    return dict_df_component_enrichment


def top_bottom_enrichment(df_averaged, dict_components, df_top, df_bottom):
    """
    top_bottom_enrichment: creates dataframes for best and worst performing LNPs, counts and their
                            formulations
        inputs:
            df_averaged : dataframe with averaged normalized counts by cell type

        output:
            d_df_components_top : dictionary containing dataframes with enrichment analysis of top
                                    performing LNPs
            d_df_components_bottom : dictionary containing dataframes with enrichment analysis of
                                        bottom performing LNPs
    """

    d_df_components_top = get_all_enrichments(df_averaged, dict_components, df_top)
    d_df_components_bottom = get_all_enrichments(df_averaged, dict_components, df_bottom)

    return d_df_components_top, d_df_components_bottom


def df_top_and_bottom(df_averaged, x_percent, number_naked_bcs):
    """
    df_top_and_bottom: creates dataframes for best and worst performing LNPs, counts and their
                            formulations
        inputs:
            df_averaged : dataframe with averaged normalized counts by cell type
            x_percent : user specified integer to find top and bottom performing LNPs (0-100)
            number_naked_bcs : user specified number of naked barcodes used
        output:
            df_top : dataframe of top performing LNPs
            df_bottom : dataframe of bottom performing LNPs
    """
    df_top, df_bottom = top_and_bottom_percent(df_averaged, x_percent, number_naked_bcs)

    return df_top, df_bottom


def top_and_bottom_percent(df_sorted, x_percent, number_naked_bcs):
    """
    CURRENTLY NOT CHECKING IF NAKED BARCODES NOT ON BOTTOM
    top_and_bottom_percent: creates dataframes for best and worst performing LNPs, counts and their\
                            formulations
        inputs:
            df_sorted : dataframe with normalized counts sorted in descending order by specified\
                        cell type
            x_percent : user specified integer to find top and bottom performing LNPs (0-100)
            number_naked_bcs : user specified number of naked barcodes
        output:
            df_top : dataframe top performing LNPs
            df_bottom : dataframe bottom performing LNPs
    """

    total_lnp = len(df_sorted.index) - number_naked_bcs  # subtract two because of naked barcodes
    values_x_percent = math.ceil(total_lnp * (x_percent / 100))

    # gets top x percent
    df_top = df_sorted.loc[range(0, values_x_percent)]
    # gets bottom x percent
    df_bottom = df_sorted.loc[range(total_lnp - values_x_percent, total_lnp + number_naked_bcs)]

    # if "NAKED1" not in df_bottom["LNP"].to_list() and "NAKED2" not in df_bottom["LNP"].to_list():
    # raise NameError("Error: Naked barcodes not on bottom " + str(x_percent) + "% + 2!")

    return df_top, df_bottom


def get_all_enrichments(df, dict_components, df_top_bottom_sort_by=None):
    """
    get_all_enrichments: calculated enrichment by component or component_ratio
        inputs:
            df : dataframe
            df_top_bottom_sort_by : dataframe of either top or bottom performing LNPs by specified
                                    sort_by (optional input)
        output:
            dict_df_components : dictionary with all dataframes of all enrichment calculations of
                                df_averaged (or df_top_bottom_sort_by if inputted)
    """
    dict_df_components = {}
    try:
        for component in dict_components:
            temp_list = dict_components[component]
            while "TOTAL" in temp_list:
                temp_list.remove("TOTAL")
    except ValueError:
        pass

    if df_top_bottom_sort_by is None:
        for component in dict_df_components:
            dict_df_components[component] = calculate_enrichment(component, dict_components[component], df)
    else:
        for component in dict_components:
            dict_df_components[component] = calculate_enrichment(component, dict_components[component],
                                                                 df_top_bottom_sort_by)

    return dict_df_components


def calculate_enrichment(component, component_list, df):
    """
    calculate_enrichment: calculated enrichment by component or component_ratio
    """

    component_total = [0] * len(component_list)
    for bc_x in df[component].values:
        for index, value in enumerate(component_list):
            if bc_x == value:
                component_total[index] += 1
                break

    total = sum(component_total)

    component_percent_total = []
    for each_component in component_total:
        component_percent_total.append(round(each_component / total, 9))

    component_list.append("TOTAL")
    component_total.append(total)
    component_percent_total.append(round(sum(component_percent_total)))

    t_component_list = [component_list, component_total, component_percent_total]
    np_temporary = np.array(t_component_list)
    np_temporary = np_temporary.T
    df_component_list = pd.DataFrame(data=np_temporary, columns=[component, "Total #",
                                                                 "% of Total"])

    return df_component_list


def get_lists_of_components(df_formulations, list_components, number_naked_bcs):
    """
    get_lists_of_components : returns a dictionary with all component mole ratios and component types
        inputs:
            df_formulations : dataframe of formulations
            list_components : list of the components used to formulate LNPs
            number_naked_bcs : user specified number of naked barcodes
        output:
            dict_components : a dictionary containing list of all the component mole ratios
                            and types
    """

    dict_components = get_dict_components(list_components)

    for component in dict_components:
        dict_components[component] = retrieve_component_list(df_formulations, component, number_naked_bcs)

    return dict_components


def get_dict_components(list_components):
    """
    get_dict_components : creates a dictionary with empty lists for each component in formulation sheet
        input :
            list_components : list of the components used to formulate LNPs
        output :
            dict_components : creates dictionary with empty list to save each type of component used
    """
    dict_components = {}

    for __, item in enumerate(list_components):
        dict_components[item] = []

    return dict_components


def retrieve_component_list(df_formulations, component, number_naked_bcs):
    """
    retrieve_component_list : returns a list of all the different mole ratios or types of a specific
                            component used
        inputs:
            df_formulations : dataframe of formulations
            component : string of the component in question
        output:
            component_list : list of all the different mole ratios or types of a component used
    """
    component_list = []

    for index in range(len(df_formulations[component].values) - number_naked_bcs):
        if df_formulations[component].values[index] not in component_list:
            component_list.append(df_formulations[component].values[index])

    component_list.sort()

    return component_list


def sort_norm_counts(df, col_num):
    """
    sort_norm_counts : sorts dataframe in descending order of norm counts
        inputs :
            df : dataframe
            col_num : column number to sort by
        output :
            df_sorted : sorted dataframe
    """
    temp_list = df.columns.tolist()
    sort_by = temp_list[col_num]
    df_sorted = df.sort_values(by=sort_by, ascending=False, ignore_index=True)
    return df_sorted


def get_df_overall(dict_df_organs, df_formulations):
    """
    get_df_overall : creates dataframe with overall average
        inputs :
            dict_df_organs : dictionary containing dataframes of all organs
            df_formulations : dataframe with formulations sheet
        output :
            df_overall : dataframe with overall average
    """

    df = pd.DataFrame()

    for key in dict_df_organs:
        df_temp = pd.concat([df, dict_df_organs[key][key + "-AVG"]], axis=1)
        df = df_temp

    avg = df.mean(axis=1)
    df_overall = pd.concat([df_formulations, df], axis=1)
    df_overall["Overall-AVG"] = avg

    return df_overall


def df_by_organs(df_merged, sorted_cells, dict_df_avg_cell_type, list_organs):
    """
    df_by_organs : creates dictionary with dataframes for all organs
        inputs :
            df_merged : dataframe containing formulation information and normalized counts
            sorted_cells : user specified list of the sorted cell types
            dict_df_avg_cell_type : dictionary with averaged dataframes of each cell type
        output :
            dict_df_organs : dictionary containing dataframes of all organs
    """

    dict_cells_by_organs = get_dict_cells_organs(sorted_cells, list_organs)
    dict_df_organs = {}

    for organ in list_organs:
        list_cells_by_organ = dict_cells_by_organs[organ]
        if len(list_cells_by_organ) == 1:
            df_lnp = df_merged["LNP"].to_frame()
            df = dict_df_avg_cell_type[list_cells_by_organ[0]]
            df = df.rename(columns={list_cells_by_organ[0]: organ + "-AVG"})
            df = df.drop(['std'], axis=1)
            dict_df_organs[organ] = pd.concat([df_lnp, df], axis=1)
        else:
            dict_df_organs[organ] = build_df_organ(df_merged, dict_df_avg_cell_type, list_cells_by_organ, organ)

    return dict_df_organs


def build_df_organ(df_merged, dict_df_avg_cell_type, list_cells_by_organ, organ):
    """
    build_df_organ : builder of dataframe for organ
        inputs :
            df_merged : dataframe containing formulation information and normalized counts
            dict_df_avg_cell_type : dictionary with averaged dataframes of each cell type
            list_cells_by_organ : list_cells_by_organs : list of cell types sorted of an organ
            organ : organ for which we wish to get dataframe
        output :
            df_organ : dataframe with data for specific organ
    """
    df_organ = df_merged["LNP"].to_frame()
    for cell_type in list_cells_by_organ:
        temp_df = dict_df_avg_cell_type[cell_type]
        df = pd.concat([df_organ, temp_df[cell_type]], axis=1)
        df_organ = df

    # get average of organ
    avg = df_organ.mean(axis=1)
    df_organ[organ + "-AVG"] = avg

    return df_organ


def get_dict_cells_organs(sorted_cells, list_organs):
    """
    get_dict_cells_organs : creates a dictionary containing cell types sorted by organ
        inputs :
            sorted_cells : user specified list of the sorted cell types
            list_organs : list of organs sorted
        output :
            dict_cells_by_organs : dictionary of cell types organized by organ
    """

    dict_cells_by_organs = {}

    for organ in list_organs:
        dict_cells_by_organs[organ] = get_list_cells_by_organ(sorted_cells, organ)

    return dict_cells_by_organs


def get_list_cells_by_organ(sorted_cells, organ):
    """
    get_list_cells_by_organ : gets list of cell types sorted for a specific organ
        inputs :
            sorted_cells : user specified list of the sorted cell types
            organ : specific organ for which we want to get the list of cell types sorted
        outputs :
            list_cells_by_organs : list of cell types sorted of an organ
    """

    list_cells_by_organ = [cell_type for cell_type in sorted_cells if cell_type[0] == organ]

    return list_cells_by_organ


def get_list_organs(sorted_cells):
    """
    get_list_organs : returns a list of the organs sorted
        input :
            sorted_cells :  user specified list of the sorted cell types
        output :
            list_organs : list of organs sorted
    """

    list_organs = []
    for cell_type in sorted_cells:
        if cell_type[0] not in list_organs:
            list_organs.append(cell_type[0])

    return list_organs


def df_cell_types(df_merged, list_samples_by_cell_type):
    """
    df_cell_types: gets dataframe of each cell type
        inputs :
            df_merged : dataframe containing formulation information and normalized counts
            list_samples_by_cell_type : lists of samples IDs by sorted cell type
        output :
            dict_df_avg_cell_type : dictionary with averaged dataframes of each cell type

    df columns titles like: LNP#   Sample1   Sample2   SampleN   Average   Stdev
    """
    dict_df_avg_cell_type = avg_cell_type(df_merged, list_samples_by_cell_type)
    df1 = df_merged["LNP"].to_frame()
    for key, value in dict_df_avg_cell_type.items():
        df = pd.concat([df1, value], axis=1)
        value.loc[:, key] = df

    return dict_df_avg_cell_type


def avg_cell_type(df_merged, dict_samples_by_cell_type):
    """
    avg_cell_type : calculates the average of each cell type
        inputs :
            df_merged : dataframe containing formulation information and normalized counts
            dict_samples_by_cell_type : dictionary containing lists of samples IDs by sorted cell type
        outputs :
            dict_df_by_cell_type : dictionary containing dataframes for each sorted cell type and its average
            and standard deviation
    """
    dict_df_by_cell_type = {}

    for cell_type in dict_samples_by_cell_type:
        dict_df_by_cell_type[cell_type] = get_df_cell_type(df_merged, dict_samples_by_cell_type[cell_type])

    for cell_type, df_cell_type in dict_df_by_cell_type.items():
        avg = df_cell_type.mean(axis=1)
        std = df_cell_type.std(axis=1)
        df_cell_type.loc[:, cell_type] = avg
        df_cell_type.loc[:, "std"] = std

    return dict_df_by_cell_type


def get_df_cell_type(df_merged, list_samples):
    """
    get_df_cell_type : returns dataframe with only samples specified
        inputs :
            df_merged : dataframe containing formulation information and normalized counts
            list_samples : list of samples from a given cell type
        outputs :
            df_cell_type : a dataframe containing all samples from a specific cell type
    """
    return df_merged[list_samples]


def divide_samples_by_cell_type(df_merged, sorted_cells):
    """
    divide_samples_by_cell_type : creates a dictionary containing cell types as keys and a list of sample IDs as the
                                    value
        inputs :
            df_merged : dataframe containing formulation information and normalized counts
            sorted_cells : user specified list of the sorted cell types
        output :
            dict_samples_by_cell_type : dictionary containing lists of samples IDs by sorted cell type
    """
    # will be a list containing list of samples organized by cell types
    dict_samples_by_cell_type = {}  # will have same length as sorted_cells

    columns_df_merged = df_merged.columns.tolist()
    columns_df_merged = columns_df_merged[11:]  # merged must include column for charge

    for cell_type in sorted_cells:
        samples_by_cell_type = [sample for sample in columns_df_merged if cell_type in sample]
        dict_samples_by_cell_type[cell_type] = samples_by_cell_type

    return dict_samples_by_cell_type


def merge_formulations_and_norm_counts(df_one, df_two, destination_file='', s_name=''):  # pylint: disable=R1710
    """
    merge_formulations_and_norm_counts : merges formulation and norm count dataframes into single
    data frame and appends it to excel spreadsheet named "Formulations + Norm Counts"
        inputs :
            df_one : first dataframe containing formulations
            df_two : second dataframe containing norm counts
            destination_file : directory of the excel spreadsheet created
            s_name = name of sheet
        output :
            df_merged : dataframe containing formulation information and normalized counts
    """
    organized_columns = organize_cell_type(df_two)

    # inner merge of data frames around barcodes ("BC")
    df_merged = df_one.merge(df_two, on="BC")

    # ordered columns
    l1 = df_one.columns.tolist()  # columns up to phospholipid%
    order_columns = l1 + organized_columns  # formulation columns and organized sample columns

    # rearrange columns on df_merged
    df_merged = df_merged[order_columns]

    if destination_file != '':
        # append merged data frames onto spreadsheet on sheet named Formulations + Norm Counts
        # with outliers
        with pd.ExcelWriter(destination_file, engine="openpyxl", mode="w") as writer:
            df_merged.to_excel(writer, sheet_name=s_name, index=False)

    return df_merged


def organize_cell_type(df_norm_counts):
    """
    organize_cell_type : takes in a dataframe and organizes the samples alphabetically
        inputs :
            df_norm_counts :  dataframe of normalized counts
        output :
            list_organized_col : list of samples organized alphabetically
    """

    list_organized_col = get_columns(df_norm_counts)

    list_organized_col.sort()
    return list_organized_col


def get_columns(dataframe):
    """
    get_columns : gets dataframe, removes barcodes and returns list of the names of columns
        inputs :
                dataframe :  a dataframe
        output :
                sample_columns : list of the names of the columns on the dataframe (names of samples)
    """

    # drop first column (barcodes)
    dataframe = dataframe.drop("BC", axis=1)

    # get row names and save barcode column
    sample_columns = dataframe.columns.tolist()

    return sample_columns


def create_df_norm_counts(csv_filepath, sample_numbers):
    """
    create_df_norm_counts : gets csv file path with normalized counts, creates a dataframe
        inputs :
            csv_filepath : file path to csv file
            sample_numbers : numbers with sample values for an experiment
        output :
            df_norm_counts : dataframe with normalized counts
    """

    # Read CSV file and save as dataframe
    df_norm_counts = pd.read_csv(csv_filepath, sep=',', header=0)

    columns = df_norm_counts.columns.tolist()  # get names of columns

    new_columns = [columns[0]]
    for i in range(1, len(columns)):
        for sample_num in sample_numbers:
            if sample_num in columns[i]:
                new_columns.append(columns[i])
                break

    df_norm_counts = df_norm_counts[new_columns]

    # rename first column to BC for barcodes
    df_norm_counts.rename(columns={columns[0]: "BC"}, inplace=True)

    return df_norm_counts


def create_df_formulation_sheet(formulations_sheet):
    """
    create_df_formulation_sheet : gets formulation sheet, creates a dataframe for formulations
        inputs :
            formulations_sheet : file path to excel sheet of formulation sheet
        output :
            df_formulations : dataframe with formulations sheet
    """

    # Turn formulation sheet into dataframe
    df_formulations = pd.read_excel(formulations_sheet, sheet_name="Formulations", header=0)

    columns = df_formulations.columns.tolist()
    df_formulations.rename(columns={columns[0]: "LNP", columns[1]: "BC"}, inplace=True)

    return df_formulations


def create_excel_spreadsheet(destination_folder, file_name="Whole Enrichment Analysis"):
    """
    create_excel_spreadsheet : creates an excel spreadsheet
        inputs :
            destination_folder : directory of the folder where the user wants the file stored
            file_name : name of the file being created (default = "Whole Enrichment Analysis")
        output :
            destination_file : directory of the excel spreadsheet created
    """
    if destination_folder[-1] != '/':  # check to save file on correct folder
        destination_folder = destination_folder + '/'

    destination_file = destination_folder + file_name + ".xlsx"
    w_b = Workbook()
    w_b.save(destination_file)

    return destination_file


if __name__ == "__main__":
    main()
